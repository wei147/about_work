import traceback
import numpy as np
import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import math
import shutil
import xlwings as xw
import time
from datetime import datetime

# 固定的“市场”字段值
MARKET_FIELD = "市场"
WAREHOUSE_FIELD = "仓区/市场"

# 背景色定义
RED_FILL = PatternFill(start_color="f07c82", end_color="f07c82", fill_type="solid")
GREEN_FILL = PatternFill(start_color="61ac85", end_color="61ac85", fill_type="solid")
DEFAULT_FILL = PatternFill(fill_type=None)


# 加载配置文件
def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# 提取平台名称
def extract_platform_name(file_name, platform_mapping):
    base_name = os.path.basename(file_name).split('-')[0]
    return base_name if base_name in platform_mapping else None


# 日志记录
def log_update_info(sku_value, custom_sku, warehouse_market, inventory_value):
    print(f"SKU: {sku_value} -> 自定义SKU: {custom_sku} -> 仓区: {warehouse_market} -> 可用库存: {inventory_value}")


# 更新库存数据的函数（共享库存）
def update_inventory_shared(ws, sku_mapping_df, shared_inventory_df, sheet_info, platform_name, source_mark):
    inventory_threshold = 20  # 默认小于20调为0
    special_platforms = {"Allegro": 10, "TEMU": 10}  # 特殊平台的库存阈值

    # 检查当前平台是否有特殊的库存阈值
    if platform_name in special_platforms:
        inventory_threshold = special_platforms[platform_name]

    # 获取SKU列和库存列的配置信息
    sku_column = sheet_info['sku_column']
    quantity_column = sheet_info.get('quantity_column', None)  # 一些平台可能没有单一的库存列
    data_start_row = sheet_info.get('data_start_row', 0)

    # 定位表头行，openpyxl从1开始计数
    header_row = sheet_info['header'] + 1
    sku_column_index = None
    quantity_column_index = None

    # 获取SKU列和库存列在工作表中的索引位置
    for cell in ws[header_row]:
        if cell.value == sku_column:
            sku_column_index = cell.column
        if quantity_column and cell.value == quantity_column:
            quantity_column_index = cell.column

    # 如果没有找到对应的SKU列，返回False
    if sku_column_index is None:
        return False

    # 遍历数据行，更新库存
    for row in ws.iter_rows(min_row=header_row + data_start_row + 1):
        sku_value = row[sku_column_index - 1].value  # 原始读取的值

        # 优先判断 None 和空字符串，不进行转换
        if sku_value is None or sku_value == "":
            continue

        # 如果是数值类型，进一步检查是否是 NaN
        if isinstance(sku_value, float) and math.isnan(sku_value):
            continue

        # 将数值和字符串都转换为字符串进行后续处理
        sku_value = str(sku_value).strip()  # 去除可能的多余空格

        quantity_cell = row[quantity_column_index - 1] if quantity_column_index else None

        # 首先匹配常规库存，获取自定义 SKU
        custom_sku = None  # 先初始化custom_sku为None，确保后续引用安全
        key = (sku_value, platform_name, source_mark)

        try:
            mapping_row = sku_mapping_df.loc[key]
            custom_sku = mapping_row['自定义sku']
            warehouse_market = mapping_row[WAREHOUSE_FIELD]
        except KeyError:
            print(f"未找到 SKU 映射，SKU: {sku_value}")
            custom_sku = None
            warehouse_market = None

        if isinstance(custom_sku, pd.Series):
            if len(custom_sku) > 1:
                print(f"警告：{source_mark}的SKU {sku_value} 的 custom_sku{custom_sku} 存在多个值，可能存在数据问题")
            # 如果 custom_sku 是 Series，取第一个值
            custom_sku = custom_sku.iloc[0]

        if isinstance(warehouse_market, pd.Series):
            warehouse_market = warehouse_market.iloc[0]

            # 根据自定义 SKU 和市场信息在共享库存表中查找对应的库存数据
            inventory_value = shared_inventory_df[
                (shared_inventory_df['SKU'] == custom_sku) &
                (shared_inventory_df[MARKET_FIELD] == warehouse_market)
                ]['可用库存'].values

            if inventory_value.size > 0:
                # 库存找到，处理
                if 0 < inventory_value[0] < inventory_threshold:  # 使用对应平台的库存阈值
                    if quantity_cell:
                        quantity_cell.value = 0
                        quantity_cell.fill = GREEN_FILL
                else:
                    if quantity_cell:
                        quantity_cell.value = inventory_value[0]
                log_update_info(sku_value, custom_sku, warehouse_market,
                                quantity_cell.value if quantity_cell else "N/A")
                continue  # 库存匹配成功，跳过当前行的其他判断
            else:
                # 没有找到库存，继续进入特殊处理逻辑
                print(f"常规库存匹配失败，继续进入特殊处理逻辑，SKU: {sku_value}")

        # 进入特殊处理逻辑，如果没有在常规匹配中找到库存
        if custom_sku is not None:
            # 特殊情况1：检查是否为组合 SKU（包含加号 "+"）
            if any(separator in custom_sku for separator in ['+']):
                # 分割字符串并去除每个 SKU 的两端空格
                combined_skus = [sku.strip() for sku in re.split(r'[+,，]', custom_sku)]
                min_inventory = float('inf')
                all_matched = True  # 标记是否所有SKU都匹配到库存

                for part_sku in combined_skus:
                    inventory_value = shared_inventory_df[
                        (shared_inventory_df['SKU'] == part_sku) &
                        (shared_inventory_df[MARKET_FIELD] == warehouse_market)
                        ]['可用库存'].values

                    if inventory_value.size > 0:
                        min_inventory = min(min_inventory, inventory_value[0])
                    else:
                        min_inventory = 0  # 如果部分SKU未匹配到库存，将库存设置为0
                        all_matched = False

                # 根据匹配结果设置库存值和背景色
                if quantity_cell:
                    quantity_cell.value = min_inventory
                    if all_matched:
                        if min_inventory == 0:
                            quantity_cell.fill = GREEN_FILL  # 所有SKU库存为0时标记为绿色
                        elif 0 < min_inventory < inventory_threshold:
                            quantity_cell.value = 0  # 库存小于阈值时将库存设为0
                            quantity_cell.fill = GREEN_FILL  # 标记为绿色
                        else:
                            quantity_cell.fill = DEFAULT_FILL  # 清除背景色
                    else:
                        quantity_cell.value = 0  # 有SKU未匹配到库存，库存设置为0
                        quantity_cell.fill = RED_FILL  # 标记为红色

                log_update_info(sku_value, '组合SKU', warehouse_market, quantity_cell.value if quantity_cell else "N/A")
                continue

            # 特殊情况2：匹配倍数标识符（* 或 x）中的自定义 SKU
            multiplier_pattern = re.compile(r'([*xX×-])(\d+)(pcs)?$', re.IGNORECASE)
            match = multiplier_pattern.search(custom_sku)
            if match:
                print(f"处理倍数SKU: {custom_sku}")  # 打印倍数SKU调试信息
                base_sku = custom_sku[:match.start()]  # 获取基础 SKU
                multiplier = int(match.group(2))  # 获取倍数

                inventory_value = shared_inventory_df[
                    (shared_inventory_df['SKU'] == base_sku) &
                    (shared_inventory_df[MARKET_FIELD] == warehouse_market)
                    ]['可用库存'].values

                if inventory_value.size > 0:
                    adjusted_inventory = math.floor(inventory_value[0] / multiplier)  # 向下取整
                    if quantity_cell:
                        quantity_cell.value = adjusted_inventory

                        # 根据库存值设置背景色   注：倍数sku是小于10才调0
                        if 0 < adjusted_inventory < 10:
                            quantity_cell.value = 0
                            quantity_cell.fill = GREEN_FILL  # 库存小于10，标记为绿色
                        else:
                            quantity_cell.fill = DEFAULT_FILL  # 清除背景色
                else:
                    if quantity_cell:
                        quantity_cell.value = 0
                        quantity_cell.fill = RED_FILL  # 匹配不到库存，标记为红色

                log_update_info(sku_value, base_sku, warehouse_market, quantity_cell.value if quantity_cell else "N/A")
                continue

        # 如果上述情况均未命中，默认将库存设置为0，并标记为红色
        if quantity_cell:
            quantity_cell.value = 0
            quantity_cell.fill = RED_FILL
        print(f"未找到匹配的SKU: {sku_value} 在 SKU映射表中")

    return True  # 如果处理成功，返回True


def convert_warehouse_ids(warehouse_ids, warehouse_mapping_df):
    """
    将平台仓库ID转换为仓库名称。

    Args:
        warehouse_ids (list): 包含平台仓库ID的列表。
        warehouse_mapping_df (DataFrame): 包含 '平台仓库ID' 和 '仓库名称' 列的仓库映射表。

    Returns:
        list: 对应的仓库名称列表。
    """
    converted_names = []
    for w_id in warehouse_ids:
        if pd.isna(w_id) or w_id == '':
            converted_names.append(None)
            continue
        # 转换为字符串以确保匹配
        w_id_str = str(w_id).strip()
        matching = warehouse_mapping_df[warehouse_mapping_df['平台仓库ID'].astype(str) == w_id_str]['仓库名称']
        if not matching.empty:
            converted_names.append(matching.values[0])
        else:
            print(f"未找到平台仓库ID: {w_id} 的对应名称")
            converted_names.append(None)
    return converted_names


def convert_warehouse_info(warehouse_infos, warehouse_mapping_df):
    """
    将平台仓库名称转换为仓库名称。

    Args:
        warehouse_infos (list): 包含平台仓库名称的列表。
        warehouse_mapping_df (DataFrame): 包含 '平台仓库名称' 和 '仓库名称' 列的仓库映射表。

    Returns:
        list: 对应的仓库名称列表。
    """
    converted_names = []
    for w_info in warehouse_infos:
        if pd.isna(w_info) or w_info == '':
            converted_names.append(None)
            continue
        # 转换为字符串以确保匹配
        w_id_str = str(w_info).strip()
        matching = warehouse_mapping_df[warehouse_mapping_df['平台仓库名称'].astype(str) == w_id_str]['仓库名称']
        if not matching.empty:
            converted_names.append(matching.values[0])
        else:
            print(f"未找到平台仓库名称: {w_info} 的对应名称")
            converted_names.append(None)
    return converted_names


def distribute_locked_inventory(locked_inventory, participating_warehouses, realtime_stocks):
    """
    分摊锁定库存。

    Args:
        locked_inventory (int): 锁定库存总数。
        participating_warehouses (list): 参与分摊的仓库名称列表。
        realtime_stocks (list): 参与分摊的仓库的实时库存列表。

    Returns:
        list: 每个仓库分摊的锁定库存列表。
    """
    num_warehouses = len(participating_warehouses)
    allocated_stocks = [0] * num_warehouses

    if num_warehouses == 0 or locked_inventory == 0:
        return allocated_stocks

    warehouses_with_stock = [
        (i, realtime_stocks[i]) for i in range(num_warehouses) if realtime_stocks[i] > 0
    ]
    num_warehouses_with_stock = len(warehouses_with_stock)

    if num_warehouses_with_stock == 0:
        return allocated_stocks

    # 初始尝试平均分摊
    if num_warehouses_with_stock <= locked_inventory:
        base_allocation = locked_inventory // num_warehouses_with_stock
        remainder = locked_inventory % num_warehouses_with_stock

        for idx, (i, stock) in enumerate(warehouses_with_stock):
            allocation = min(base_allocation, stock)
            allocated_stocks[i] = allocation

        # 分配余数部分
        remainder_indices = sorted(
            warehouses_with_stock,
            key=lambda x: x[1],  # 按实时库存排序
            reverse=True
        )
        remainder_count = remainder
        for i, stock in remainder_indices:
            if remainder_count == 0:
                break
            available_stock = stock - allocated_stocks[i]
            if available_stock > 0:
                allocated_stocks[i] += 1
                remainder_count -= 1

    else:
        # 仓库数量大于锁定库存，无法平均分摊
        # 将锁定库存全部分配给实时库存最大的仓库
        sorted_warehouses = sorted(
            warehouses_with_stock,
            key=lambda x: x[1],  # 按实时库存排序
            reverse=True
        )
        remaining_locked_inventory = locked_inventory
        for i, stock in sorted_warehouses:
            allocation = min(remaining_locked_inventory, stock)
            allocated_stocks[i] += allocation
            remaining_locked_inventory -= allocation
            if remaining_locked_inventory == 0:
                break

    return allocated_stocks


def extract_multiplier_sku(sku_value):
    """
    从倍数SKU中提取基础SKU和倍数系数。
    :param sku_value: SKU字符串，例如 "HOGA-XK00592x2"
    :return: (base_sku, multiplier)，如果提取失败则返回 (None, None)
    """
    import re
    pattern = r'^(.*?)[xX\*](\d+)$'
    match = re.match(pattern, sku_value)
    if match:
        base_sku = match.group(1)
        multiplier = int(match.group(2))
        return base_sku, multiplier
    else:
        return None, None


def extract_combination_skus(sku_value):
    """
    从组合 SKU 中提取基础 SKU 列表。
    :param sku_value: SKU 字符串，例如 "SKU1+SKU2+SKU3"
    :return: 基础 SKU 列表，例如 ["SKU1", "SKU2", "SKU3"]
    """
    if '+' in sku_value:
        base_skus = [sku.strip() for sku in sku_value.split('+')]
        return base_skus
    else:
        return None


def update_inventory_separate(ws, sku_mapping_df, realtime_inventory_df, warehouse_mapping_df, sheet_info,
                              platform_name, source_mark, shared_inventory_df):
    """
    更新库存数据的函数（独立库存），支持特殊平台的仓库ID和库存在同一列的情况。
    """
    inventory_threshold = 20
    special_platforms = {"Allegro": 10, "TEMU": 10}
    if platform_name in special_platforms:
        inventory_threshold = special_platforms[platform_name]

    sku_column = sheet_info['sku_column']
    data_start_row = sheet_info.get('data_start_row', 0)
    convert_warehouse_id = sheet_info.get('convert_warehouse_id', False)
    convert_platform_warehouse_name = sheet_info.get('convert_platform_warehouse_name', False)
    summary_column = sheet_info.get('summary_column', None)

    warehouse_pairs = sheet_info.get('warehouse_pairs', [])
    header_row = sheet_info['header'] + 1
    sku_column_index = None

    # 新增变量  特殊平台的仓库ID和库存在同一列的情况。
    combined_warehouse_inventory_column = sheet_info.get('combined_warehouse_inventory_column', False)

    # 获取SKU列的索引
    for cell in ws[header_row]:
        if cell.value == sku_column:
            sku_column_index = cell.column
            break

    if sku_column_index is None:
        print(f"未找到SKU列: {sku_column} 在表 {ws.title} 中")
        return False

    warehouse_stock_indices = []
    warehouse_id_indices = []
    warehouse_ids = []

    if platform_name == "TikTok":
        # 对于TK平台，从列标题中提取仓库ID
        for cell in ws[header_row]:
            cell_value = str(cell.value)
            # 使用固定的模式提取仓库ID
            if cell_value.startswith("Quantity in "):
                warehouse_id = cell_value[len("Quantity in "):].strip()
                warehouse_stock_indices.append(cell.column - 1)
                warehouse_ids.append(warehouse_id)
    else:
        for pair in warehouse_pairs:
            id_col = pair.get('name_column')
            stock_col = pair['stock_column']
            id_index = None
            stock_index = None
            for cell in ws[header_row]:
                if id_col and cell.value == id_col:
                    id_index = cell.column - 1
                if cell.value == stock_col:
                    stock_index = cell.column - 1
            if stock_index is not None:
                warehouse_stock_indices.append(stock_index)
                if id_index is not None:
                    warehouse_id_indices.append(id_index)
                else:
                    warehouse_id_indices.append(None)
            else:
                print(f"未找到库存列: {stock_col}")
                return False

    summary_column_index = None
    if summary_column:
        for cell in ws[header_row]:
            if cell.value == summary_column:
                summary_column_index = cell.column - 1
                break
        if summary_column_index is None:
            print(f"未找到汇总列: {summary_column} 在表 {ws.title} 中")
            return False

    market_mapping = {
        '美国仓': '美',
        '欧洲仓': '德',
        '英国仓': '英',
        '加拿大仓': '加',
        '澳洲仓': '澳'
    }

    for ws_row in ws.iter_rows(min_row=header_row + data_start_row + 1):
        sku_value = ws_row[sku_column_index - 1].value

        # 优先判断 None 和空字符串，不进行转换
        if sku_value is None or sku_value == "":
            continue

        # 如果是数值类型，进一步检查是否是 NaN
        if isinstance(sku_value, float) and math.isnan(sku_value):
            continue

        # 将数值和字符串都转换为字符串进行后续处理
        sku_value = str(sku_value).strip()  # 去除可能的多余空格
        print(sku_value)

        total_updated_inventory = 0
        custom_sku = None
        warehouse_market = None

        key = (sku_value, platform_name, source_mark)
        try:
            mapping_row = sku_mapping_df.loc[key]
            custom_sku = mapping_row['自定义sku']
            warehouse_market = mapping_row[WAREHOUSE_FIELD]
        except KeyError:
            print(f"未找到 SKU 映射，SKU: {sku_value}")
            custom_sku = None
            warehouse_market = None

        if isinstance(custom_sku, pd.Series):
            if len(custom_sku) > 1:
                print(f"警告：{source_mark}的SKU {sku_value} 的 custom_sku{custom_sku} 存在多个值，可能存在数据问题")
            # 如果 custom_sku 是 Series，取第一个值
            custom_sku = custom_sku.iloc[0]

        if isinstance(warehouse_market, pd.Series):
            warehouse_market = warehouse_market.iloc[0]

        if custom_sku is None or warehouse_market is None:
            for stock_idx in warehouse_stock_indices:
                ws_row[stock_idx].value = 0
                ws_row[stock_idx].fill = RED_FILL
            log_update_info(sku_value, "未匹配", "N/A", 0)
            if summary_column and summary_column_index is not None:
                ws_row[summary_column_index].value = 0
            print("-" * 50)
            continue

        market_keyword = market_mapping.get(warehouse_market, '')
        if not market_keyword:
            for stock_idx in warehouse_stock_indices:
                ws_row[stock_idx].value = 0
                ws_row[stock_idx].fill = RED_FILL
            if summary_column and summary_column_index is not None:
                ws_row[summary_column_index].value = 0
            print("匹配实时库存表中仓库名称失败")
            print("-" * 50)
            continue

        inventory_rows = realtime_inventory_df[realtime_inventory_df['自定义SKU'] == custom_sku]
        # 没有找到对应的库存行，只有表头。
        # if inventory_rows.empty:
        #     for stock_idx in warehouse_stock_indices:
        #         ws_row[stock_idx].value = 0
        #         ws_row[stock_idx].fill = RED_FILL
        #     log_update_info(sku_value, custom_sku, "没有找到对应的库存行，只有表头", 0)
        #     if summary_column and summary_column_index is not None:
        #         ws_row[summary_column_index].value = 0
        #     print("-" * 50)
        #     continue

        locked_inventory_rows = shared_inventory_df[
            (shared_inventory_df['SKU'] == custom_sku) &
            (shared_inventory_df[MARKET_FIELD] == warehouse_market)
            ]

        # **添加倍数SKU处理逻辑开始**
        if locked_inventory_rows.empty:
            # 检查是否为组合 SKU
            # if custom_sku =="HOGA-XK00327-A+HOGA-XK00327-B":
            #     a=1
            base_skus = extract_combination_skus(custom_sku)
            if base_skus:
                print(f"检测到组合 SKU，基础 SKU 列表: {base_skus}")
                # 获取仓库名称列表
                if platform_name == "TikTok":
                    warehouse_names_in_row = warehouse_ids.copy()
                else:
                    warehouse_ids_in_row = [ws_row[idx].value if idx is not None else None for idx in
                                            warehouse_id_indices]
                    warehouse_ids_in_row = [str(name).strip() if name is not None else None for name in
                                            warehouse_ids_in_row]
                    warehouse_names_in_row = warehouse_ids_in_row.copy()

                # 转换仓库ID为仓库名称
                if convert_warehouse_id:
                    warehouse_names_in_row = convert_warehouse_ids(warehouse_ids_in_row, warehouse_mapping_df)

                # 转换平台仓库名称为仓库名称。其中因为TK平台比较特殊，库存列和仓库列在同一列
                if convert_platform_warehouse_name:
                    if platform_name == "TikTok":
                        warehouse_names_in_row = convert_warehouse_info(warehouse_names_in_row, warehouse_mapping_df)
                    else:
                        warehouse_names_in_row = convert_warehouse_info(warehouse_ids_in_row, warehouse_mapping_df)

                # 初始化一个字典来存储每个仓库的可用库存
                warehouse_combination_stocks = {warehouse_name: None for warehouse_name in warehouse_names_in_row}

                # 遍历每个基础 SKU
                for base_sku in base_skus:
                    print(f"处理基础 SKU: {base_sku}")
                    warehouse_results = []  # 用于存储仓库处理结果
                    # 获取基础 SKU 的实时库存和锁定库存
                    inventory_rows = realtime_inventory_df[realtime_inventory_df['自定义SKU'] == base_sku]
                    locked_inventory_rows = shared_inventory_df[
                        (shared_inventory_df['SKU'] == base_sku) &
                        (shared_inventory_df[MARKET_FIELD] == warehouse_market)
                        ]
                    if inventory_rows.empty or locked_inventory_rows.empty:
                        print(f"未找到基础 SKU 的库存数据，基础 SKU: {base_sku}")
                        # 对于所有仓库，将库存设为 0，背景色设为红色
                        for stock_idx in warehouse_stock_indices:
                            ws_row[stock_idx].value = 0
                            ws_row[stock_idx].fill = RED_FILL  # 仅在数据缺失时设置红色背景
                        if summary_column and summary_column_index is not None:
                            ws_row[summary_column_index].value = 0
                        print("-" * 50)
                        break  # 退出基础 SKU 的循环，因为数据缺失
                    else:
                        # 获取锁定库存
                        locked_inventory = locked_inventory_rows.iloc[0]['锁定库存']
                        # 收集所有有库存的仓库（从实时库存表中获取）
                        all_warehouse_columns = inventory_rows.columns.tolist()
                        participating_warehouses = []
                        realtime_stocks = []
                        for warehouse_name in all_warehouse_columns:
                            if market_keyword in warehouse_name:
                                original_stock = inventory_rows.iloc[0][warehouse_name]
                                if original_stock > 0:
                                    participating_warehouses.append(warehouse_name)
                                    realtime_stocks.append(original_stock)

                        # 检查参与分摊的仓库是否包含 Excel 表中的仓库名称
                        common_warehouses = set(participating_warehouses) & set(warehouse_names_in_row)

                        if not participating_warehouses or not common_warehouses:
                            # 没有参与分摊的仓库，或者参与分摊的仓库不在 Excel 表中
                            # 对于 Excel 表中的每个仓库，检查其是否存在于库存数据中
                            for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
                                if name_idx in all_warehouse_columns:
                                    # 仓库存在于库存数据中，但库存为 0
                                    warehouse_combination_stocks[name_idx] = 0
                                    # 背景色保持默认
                                else:
                                    # 仓库不存在于库存数据中
                                    warehouse_combination_stocks[name_idx] = {'value': 0, 'red_fill': True}
                            print(f"基础 SKU {base_sku} 没有可用的参与分摊仓库")
                            continue  # 处理下一个基础 SKU

                        # 分摊锁定库存
                        allocated_locked_inventory = distribute_locked_inventory(
                            locked_inventory, participating_warehouses, realtime_stocks
                        )

                        # 计算更新后的库存
                        for idx, warehouse_name in enumerate(participating_warehouses):
                            original_stock = inventory_rows.iloc[0][warehouse_name]
                            allocation = allocated_locked_inventory[idx]
                            updated_stock = original_stock - allocation
                            if updated_stock < 0:
                                updated_stock = 0
                            # 如果仓库在 Excel 表中，需要更新最小库存
                            if warehouse_name in warehouse_combination_stocks:
                                if warehouse_combination_stocks[warehouse_name] is None:
                                    warehouse_combination_stocks[warehouse_name] = updated_stock
                                else:
                                    warehouse_combination_stocks[warehouse_name] = min(
                                        warehouse_combination_stocks[warehouse_name], updated_stock
                                    )
                            # 添加仓库处理结果到列表
                            warehouse_results.append(
                                f"- {warehouse_name}: 原始库存={original_stock}, 分摊锁定库存={allocation}, 更新后库存={updated_stock}"
                            )
                            # 在处理完该基础 SKU 后，输出仓库分配结果。打印日志
                        # print("仓库分配结果:")
                        print(f"锁定库存数={locked_inventory}")
                        for result in warehouse_results:
                            print(result)
                        # 对于不在参与分摊的仓库，如果在 Excel 表中，且未设置库存，则设为 0
                        for name_idx in warehouse_combination_stocks.keys():
                            if name_idx not in participating_warehouses and warehouse_combination_stocks[
                                name_idx] is None:
                                warehouse_combination_stocks[name_idx] = 0
                else:
                    # 初始化总更新库存量
                    total_updated_inventory = 0

                    # 遍历仓库名称和对应的库存列索引
                    for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
                        # 获取组合库存信息
                        combination_stock_info = warehouse_combination_stocks.get(name_idx, None)
                        combination_stock = None  # 初始化库存值

                        if combination_stock_info is None:
                            # 数据缺失，设置库存为 0，背景色为红色
                            ws_row[stock_idx].value = 0
                            ws_row[stock_idx].fill = RED_FILL
                            continue  # 跳过后续处理
                        elif isinstance(combination_stock_info, dict):
                            # 包含值和背景色信息
                            combination_stock = combination_stock_info.get('value', 0)
                            ws_row[stock_idx].value = combination_stock
                            if combination_stock_info.get('red_fill'):
                                ws_row[stock_idx].fill = RED_FILL
                        else:
                            # 直接获取库存值
                            combination_stock = combination_stock_info

                        # 根据库存值进行处理
                        if combination_stock == 0:
                            ws_row[stock_idx].value = 0
                            if name_idx not in all_warehouse_columns:
                                ws_row[stock_idx].fill = RED_FILL  # 仓库不存在于库存数据中，设置红色背景
                            # 背景色保持默认
                        elif 0 < combination_stock < inventory_threshold:
                            # 库存低于阈值，设置为 0，背景色为绿色
                            combination_stock = 0
                            ws_row[stock_idx].value = 0
                            ws_row[stock_idx].fill = GREEN_FILL
                        else:
                            # 库存正常，设置库存值
                            ws_row[stock_idx].value = combination_stock
                            # 背景色保持默认

                        # 累加总库存
                        total_updated_inventory += combination_stock

                    # 更新汇总列
                    if summary_column and summary_column_index is not None:
                        ws_row[summary_column_index].value = total_updated_inventory

                    # 记录日志
                    log_update_info(sku_value, custom_sku, warehouse_market, total_updated_inventory)
                    print(f"组合 SKU {custom_sku} 的库存更新完成，合计库存: {total_updated_inventory}")
                    print("-" * 50)
                    continue  # 处理完毕，继续下一行

                    # # 更新工作表中的库存
                    # total_updated_inventory = 0
                    # for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
                    #     combination_stock_info = warehouse_combination_stocks.get(name_idx, None)
                    #     combination_stock = warehouse_combination_stocks.get(name_idx, None)
                    #     if combination_stock is None:
                    #         ws_row[stock_idx].value = 0
                    #         ws_row[stock_idx].fill = RED_FILL  # 数据缺失，设置红色背景
                    #     else:
                    #         if isinstance(combination_stock_info, dict):
                    #             # 包含值和背景色信息
                    #             ws_row[stock_idx].value = combination_stock_info['value']
                    #             if combination_stock_info.get('red_fill'):
                    #                 ws_row[stock_idx].fill = RED_FILL
                    #         else:
                    #             combination_stock = combination_stock_info
                    #             # combination_stock = combination_stock_info
                    #         if combination_stock == 0:
                    #             ws_row[stock_idx].value = 0
                    #             if name_idx not in all_warehouse_columns:
                    #                 ws_row[stock_idx].fill = RED_FILL  # 仓库不存在于库存数据中，设置红色背景
                    #             # 不改变背景色，保持默认
                    #         elif 0 < combination_stock < inventory_threshold:
                    #             # 如果库存数小于阈值，那么这里会调0。total_updated_inventory在累加的时候不会加错
                    #             combination_stock = 0
                    #             ws_row[stock_idx].value = 0
                    #             ws_row[stock_idx].fill = GREEN_FILL
                    #         else:
                    #             ws_row[stock_idx].value = combination_stock
                    #             # 不改变背景色，保持默认
                    #         total_updated_inventory += combination_stock
                    # if summary_column and summary_column_index is not None:
                    #     ws_row[summary_column_index].value = total_updated_inventory
                    # # 记录日志
                    # log_update_info(sku_value, custom_sku, warehouse_market, total_updated_inventory)
                    # print(f"组合 SKU {custom_sku} 的库存更新完成，合计库存: {total_updated_inventory}")
                    # print("-" * 50)
                    # continue  # 处理完毕，继续下一行
            # 检查是否为倍数SKU
            else:
                base_sku, multiplier = extract_multiplier_sku(custom_sku)
                if base_sku and multiplier:
                    print(f"检测到倍数 SKU，基础 SKU: {base_sku}, 倍数: {multiplier}")
                    # 使用基础 SKU 获取库存数据
                    inventory_rows = realtime_inventory_df[realtime_inventory_df['自定义SKU'] == base_sku]
                    locked_inventory_rows = shared_inventory_df[
                        (shared_inventory_df['SKU'] == base_sku) &
                        (shared_inventory_df[MARKET_FIELD] == warehouse_market)
                        ]
                    if locked_inventory_rows.empty or inventory_rows.empty:
                        print(f"未找到基础 SKU 的库存数据，基础 SKU: {base_sku}")
                        # 将库存设为 0，背景色设为红色
                        for stock_idx in warehouse_stock_indices:
                            ws_row[stock_idx].value = 0
                            ws_row[stock_idx].fill = RED_FILL
                        if summary_column and summary_column_index is not None:
                            ws_row[summary_column_index].value = 0
                        print("-" * 50)
                        continue
                    # 使用 base_sku 的库存数据，更新 custom_sku
                    custom_sku = base_sku
                else:
                    print(f"无法解析倍数 SKU，SKU: {custom_sku}")
                    # 将库存设为 0，背景色设为红色
                    for stock_idx in warehouse_stock_indices:
                        ws_row[stock_idx].value = 0
                        ws_row[stock_idx].fill = RED_FILL
                    if summary_column and summary_column_index is not None:
                        ws_row[summary_column_index].value = 0
                    print("-" * 50)
                    continue
                # **添加倍数SKU处理逻辑结束**
            # **添加倍数 SKU 和组合 SKU 处理逻辑结束**

            # 以下是库存分摊和更新逻辑（常规 SKU）

        # 没有找到对应的库存行，只有表头。
        if inventory_rows.empty:
            for stock_idx in warehouse_stock_indices:
                ws_row[stock_idx].value = 0
                ws_row[stock_idx].fill = RED_FILL
            log_update_info(sku_value, custom_sku, "没有找到对应的库存行，只有表头", 0)
            if summary_column and summary_column_index is not None:
                ws_row[summary_column_index].value = 0
            print("-" * 50)
            continue

        # 以下是库存分摊和更新逻辑
        locked_inventory = locked_inventory_rows.iloc[0]['锁定库存']
        # 注意：锁定库存不需要除以倍数系数

        participating_warehouses = []
        realtime_stocks = []

        warehouse_columns = inventory_rows.columns.tolist()
        for warehouse_name in warehouse_columns:
            # 错误发生在 3号仓-美西3仓。 也就是在实时库存表里面没有没有对应自定义sku就报错了。
            if market_keyword in warehouse_name and inventory_rows.iloc[0][warehouse_name] > 0:
                original_stock = inventory_rows.iloc[0][warehouse_name]
                # 如果是倍数SKU，需要调整实时库存
                if 'multiplier' in locals() and multiplier is not None and multiplier != 0:
                    adjusted_stock = original_stock // multiplier
                else:
                    adjusted_stock = original_stock
                participating_warehouses.append(warehouse_name)
                realtime_stocks.append(adjusted_stock)
        # 获取仓库名称列表
        if platform_name == "TikTok":
            warehouse_names_in_row = warehouse_ids.copy()
        else:
            warehouse_ids_in_row = [ws_row[idx].value if idx is not None else None for idx in
                                    warehouse_id_indices]
            warehouse_ids_in_row = [str(name).strip() if name is not None else None for name in
                                    warehouse_ids_in_row]
            warehouse_names_in_row = warehouse_ids_in_row.copy()

        if convert_warehouse_id:
            warehouse_names_in_row = convert_warehouse_ids(warehouse_ids_in_row, warehouse_mapping_df)

        # 转换平台仓库名称为仓库名称。其中因为TK平台比较特殊，库存列和仓库列在同一列
        if convert_platform_warehouse_name:
            if platform_name == "TikTok":
                warehouse_names_in_row = convert_warehouse_info(warehouse_names_in_row, warehouse_mapping_df)
            else:
                warehouse_names_in_row = convert_warehouse_info(warehouse_ids_in_row, warehouse_mapping_df)

        # 检查参与分摊的仓库是否包含 Excel 表中的仓库名称
        common_warehouses = set(participating_warehouses) & set(warehouse_names_in_row)

        if not participating_warehouses or not common_warehouses:
            # 没有参与分摊的仓库，或者参与分摊的仓库不在 Excel 表中
            # 对于 Excel 表中的每个仓库，检查其是否存在于库存数据中
            for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
                if name_idx in warehouse_columns:
                    # 仓库存在于库存数据中，但库存为 0
                    ws_row[stock_idx].value = 0
                    # 背景色保持默认
                else:
                    # 仓库不存在于库存数据中
                    ws_row[stock_idx].value = 0
                    ws_row[stock_idx].fill = RED_FILL  # 设置红色背景
            if summary_column and summary_column_index is not None:
                ws_row[summary_column_index].value = 0
            print(f"SKU：{sku_value} -> 自定义 SKU：{custom_sku} 没有可用的参与分摊仓库")
            print("-" * 50)
            continue

        allocated_locked_inventory = distribute_locked_inventory(locked_inventory, participating_warehouses,
                                                                 realtime_stocks)

        # 详细日志记录
        print(f"SKU: {sku_value}, 自定义 SKU: {custom_sku}, 仓区: {warehouse_market}, 锁定库存: {locked_inventory}")
        print(f"参与分摊的仓库: {participating_warehouses}, 分摊仓库数量: {len(participating_warehouses)}")
        print(f"对应的调整后实时库存: {realtime_stocks}")

        for warehouse_name, allocation in zip(participating_warehouses, allocated_locked_inventory):
            original_stock = inventory_rows.iloc[0][warehouse_name]
            # 计算调整后的实时库存
            if 'multiplier' in locals() and multiplier is not None and multiplier != 0:
                adjusted_stock = original_stock // multiplier
            else:
                adjusted_stock = original_stock
            # 计算更新后的库存
            updated_stock = adjusted_stock - allocation
            if updated_stock < 0:
                updated_stock = 0

            for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
                if name_idx == warehouse_name:
                    # 根据库存值设置背景色
                    if updated_stock == 0:
                        ws_row[stock_idx].value = 0
                    elif 0 < updated_stock < inventory_threshold:
                        ws_row[stock_idx].value = 0
                        ws_row[stock_idx].fill = GREEN_FILL
                        updated_stock = 0
                    else:
                        ws_row[stock_idx].value = updated_stock
                        ws_row[stock_idx].fill = DEFAULT_FILL
                    total_updated_inventory += updated_stock
                    # 记录参与分摊的仓库及其库存分配
                    print(f"原始表仓库: {warehouse_name}, 分到的锁定库存: {allocation}")
                    break

        for name_idx, stock_idx in zip(warehouse_names_in_row, warehouse_stock_indices):
            if name_idx not in participating_warehouses:
                ws_row[stock_idx].value = 0
                ws_row[stock_idx].fill = RED_FILL

        if summary_column and summary_column_index is not None:
            ws_row[summary_column_index].value = total_updated_inventory

        log_update_info(sku_value, custom_sku, warehouse_market, total_updated_inventory)
        print("-" * 50)

    return True


def convert_xls_to_xlsx(xls_file_path, output_temp_dir):
    """
      将 .xls 文件转换为 .xlsx，并存储在指定的输出目录下对应的子目录中。
      """
    # 获取原始文件的父级文件夹名称
    original_folder_name = os.path.basename(os.path.dirname(xls_file_path))

    # 生成新的 .xlsx 文件路径，存放在输出目录下的子目录中
    target_folder = os.path.join(output_temp_dir, original_folder_name)
    os.makedirs(target_folder, exist_ok=True)  # 如果子目录不存在则创建

    # 新的 .xlsx 文件路径
    base_name = os.path.basename(xls_file_path).replace('.xls', '.xlsx')
    xlsx_file_path = os.path.join(target_folder, base_name)

    # 启动 Excel 应用，隐藏窗口
    app = xw.App(visible=False)

    try:
        # 打开 .xls 文件并保存为 .xlsx
        wb = app.books.open(xls_file_path)
        wb.save(xlsx_file_path)
        print(f"已成功将 {xls_file_path} 转换为 {xlsx_file_path}")
    finally:
        wb.close()
        app.quit()

    return xlsx_file_path


def convert_csv_to_excel(csv_file_path, output_temp_dir):
    """
      将 CSV 文件转换为 .xlsx，并存储在指定的输出目录下对应的子目录中。
      """
    # 获取原始文件的父级文件夹名称
    original_folder_name = os.path.basename(os.path.dirname(csv_file_path))

    # 生成新的 .xlsx 文件路径，存放在输出目录下的子目录中
    target_folder = os.path.join(output_temp_dir, original_folder_name)
    os.makedirs(target_folder, exist_ok=True)  # 如果子目录不存在则创建

    # 新的 .xlsx 文件路径
    base_name = os.path.basename(csv_file_path).replace('.csv', '.xlsx')
    excel_file_path = os.path.join(target_folder, base_name)

    # 尝试使用不同编码读取 CSV 文件
    encodings = ['utf-8', 'utf-8-sig', 'gbk', 'latin1']
    df = None

    for encoding in encodings:
        try:
            df = pd.read_csv(csv_file_path, encoding=encoding)
            break
        except UnicodeDecodeError as e:
            print(f"读取文件时编码错误: {e}, 尝试其他编码")
        except Exception as e:
            print(f"读取文件时发生错误: {e}")
            break

    if df is not None:
        try:
            df.to_excel(excel_file_path, index=False)
            print(f"成功将 CSV 转换为 Excel: {excel_file_path}")
        except Exception as e:
            print(f"保存 Excel 文件时发生错误: {e}")
    else:
        print("未能读取 CSV 文件")

    return excel_file_path


def convert_excel_to_csv(excel_file, sheet_name):
    df = pd.read_excel(excel_file, engine='openpyxl', sheet_name=sheet_name)
    df.columns = ["" if 'Unnamed' in str(col) else col for col in df.columns]

    if excel_file.endswith(('.xlsx', '.xlsm', '.xls')):
        base_name = os.path.splitext(excel_file)[0]
    else:
        base_name = excel_file

    csv_file = f"{base_name}_{sheet_name}.csv"
    df.to_csv(csv_file, encoding='utf-8', index=False)
    print(f"CSV文件已保存至: {csv_file}")


def convert_xlsx_to_xls(xlsx_file_path):
    # 启动 Excel 应用程序，隐藏窗口
    app = xw.App(visible=False)

    try:
        # 禁用 Excel 警告窗口（如兼容性检查器）
        app.display_alerts = False
        app.screen_updating = False

        # 打开 .xlsx 文件
        wb = app.books.open(xlsx_file_path)

        # 生成新的 .xls 文件路径
        xls_file_path = xlsx_file_path.replace('.xlsx', '.xls')

        # 使用 Excel 的 SaveAs 方法将文件保存为 .xls 格式
        wb.api.SaveAs(xls_file_path, FileFormat=56)  # 56 是 Excel 97-2003 文件格式 (.xls)
        print(f"已成功将 {xlsx_file_path} 转换为 {xls_file_path}")
    finally:
        # 确保关闭工作簿和 Excel 应用程序，避免程序卡住和 Excel 残留在后台
        wb.close()
        app.quit()


# 用于把xls和csv转为 xlsx文件
def convert_file_type_if_needed(file_path, output_temp_dir):
    """
    根据文件后缀进行转换，csv -> xlsx, xls -> xlsx
    """
    if file_path.endswith('.csv'):
        file_path = convert_csv_to_excel(file_path, output_temp_dir)

    if file_path.endswith('.xls'):
        file_path = convert_xls_to_xlsx(file_path, output_temp_dir)

    return file_path


def process_inventory_file(file_path, config, sku_mapping_df, shared_inventory_df, realtime_inventory_df,
                           warehouse_mapping_df, platform_mapping, output_dir,
                           result_log, platform_config, output_temp_dir):
    print("========================================================")

    # 文件类型转换
    file_path = convert_file_type_if_needed(file_path, output_temp_dir)

    base_name = os.path.basename(file_path).split('-')[0]
    print(f"base_name: {base_name}")

    # 检查平台配置
    if platform_config is None:
        print(f"Platform name {base_name} not found in configuration.")
        result_log.append(
            f"{os.path.basename(file_path)} skipped (platform name {base_name} not found in configuration).")
        return False

    # 获取最大可以处理的sheet数量，如果未定义，则默认不限制
    max_sheets_to_process = platform_config.get("max_sheets_to_process", float('inf'))  # 未设置时默认不限制
    processed_sheets_count = 0  # 记录已处理的sheet数量
    successfully_processed_sheets = []  # 记录成功处理的 sheetname

    convert_to_csv = platform_config.get("convert_to_csv", False)
    convert_to_xls = platform_config.get("convert_to_xls", False)
    sheet_infos = platform_config['sheets_to_process']

    # 处理Excel文件
    wb = load_workbook(filename=file_path)
    sheet_visibility = {}  # 保存工作表的可见状态
    result_generated = False
    all_sheets_processed = True

    for sheet_info in sheet_infos:
        sheet_names = sheet_info.get('sheet_names', [])
        skip_if_contains = sheet_info.get('skip_if_contains', [])
        process_all = sheet_info.get('process_all', False)

        if process_all and not sheet_names:
            sheet_names = wb.sheetnames  # 如果设置了 process_all，且没有指定 sheet_names，处理所有工作表

        for sheet_name in sheet_names:
            # 如果处理的 sheet 超过了限制，跳过
            if processed_sheets_count >= max_sheets_to_process:
                print(f"Reached maximum sheet limit for {base_name}. Skipping remaining sheets.")
                break

            # 如果工作表名包含跳过的字符串，则跳过
            if any(skip_str in sheet_name for skip_str in skip_if_contains):
                print(f"Skipping sheet: {sheet_name} because it contains one of {skip_if_contains}")
                continue

            ws = wb[sheet_name]
            sheet_processed = False

            # 获取列信息
            sku_column = sheet_info.get('sku_column')
            quantity_column = sheet_info.get('quantity_column')

            print(
                f"Processing sheet: {sheet_name} with SKU column: {sku_column} and Quantity column: {quantity_column}")

            try:
                platform_name = platform_mapping.get(base_name)
                source_mark = os.path.basename(os.path.dirname(file_path))

                # 处理库存数据
                if platform_config.get('inventory_type') == 'shared':
                    if update_inventory_shared(ws, sku_mapping_df, shared_inventory_df, sheet_info, platform_name,
                                               source_mark):
                        result_generated = True
                        sheet_processed = True
                elif platform_config.get('inventory_type') == 'separate':
                    if update_inventory_separate(ws, sku_mapping_df, realtime_inventory_df, warehouse_mapping_df,
                                                 sheet_info, platform_name, source_mark, shared_inventory_df):
                        result_generated = True
                        sheet_processed = True

                # 增加已处理的 sheet 数量
                if sheet_processed:
                    processed_sheets_count += 1
                    successfully_processed_sheets.append(sheet_name)  # 记录成功处理的 sheetname

            except Exception as e:
                # 捕获异常并记录详细错误信息
                print(f"Error processing sheet {sheet_name}: {e}")
                result_log.append(f"Error processing {sheet_name} in {file_path}: {e}")
                all_sheets_processed = False

            print(f"Sheet {sheet_name} processed.")
            sheet_visibility[sheet_name] = ws.sheet_state  # 保存工作表的状态

            # 如果没有成功处理该sheet，记录未处理成功的日志
            if not sheet_processed:
                result_log.append(
                    f"{sheet_name} in {file_path} 未处理成功。检查列名配置: SKU 列为 {sku_column}, 库存列为 {quantity_column}.")
                all_sheets_processed = False

    # 保存和转换文件
    if result_generated:
        result_file = os.path.join(output_dir, os.path.basename(file_path).replace(".xlsx", "_result.xlsx"))
        wb.save(result_file)

        # 恢复工作表状态
        wb = load_workbook(result_file)
        for sheet_name, state in sheet_visibility.items():
            wb[sheet_name].sheet_state = state
        wb.save(result_file)
        print(f"处理完成，Excel结果表已保存至: {result_file}")

        # 转换为 CSV，只转换成功处理的 sheetname
        if convert_to_csv:
            for sheet_name in successfully_processed_sheets:
                convert_excel_to_csv(result_file, sheet_name)
                print(f"核心 sheet: {sheet_name} 的 CSV 结果表已生成。")

        if convert_to_xls:
            convert_xlsx_to_xls(result_file)

        if all_sheets_processed:
            result_log.append(f"{os.path.basename(file_path)} processed successfully.")
        else:
            result_log.append(f"{os.path.basename(file_path)} processed with some errors.")
        return True

    else:
        print("未生成任何结果文件，可能未找到有效的列名配置。")
        result_log.append(f"{os.path.basename(file_path)} not processed. 可能未找到有效的列名配置.")
        return False


def process_folder(folder_path, config, sku_mapping_df, shared_inventory_df, realtime_inventory_df,
                   warehouse_mapping_df, platform_mapping, output_dir, output_temp_dir):
    result_log = []
    all_files_processed_successfully = True

    # 优先级列表，优先处理.xlsx，其次.xlsm，其次.xls，最后.csv
    priority_order = ['.xlsx', '.xlsm', '.xls', '.csv']
    processed_files = set()  # 记录已经处理的文件的基础名称

    # 遍历文件夹中的文件
    for file_name in os.listdir(folder_path):
        # 跳过包含“SKU映射”的文件
        if "SKU映射" in file_name:
            print(f"Skipping file: {file_name} (matches 'SKU映射')")
            continue

        if file_name == "OS-ROSE_HGX.xlsx":
            continue

        # 获取文件的基础名称和后缀
        base_name, ext = os.path.splitext(file_name)


        # 提取平台名：平台名是文件名的第一部分
        # 假设文件名格式为 "平台名-其他部分.xlsx"
        platform_name = base_name.split('-')[0]

        # 处理优先级文件，如果同名的高优先级文件已经处理，则跳过低优先级文件
        if base_name in processed_files:
            print(f"Skipping file: {file_name} (lower priority than previously processed file)")
            continue

        # 如果后缀不在优先级列表中，则跳过该文件
        if ext not in priority_order:
            result_log.append(f"{file_name} skipped (invalid file extension).")
            continue

        # 找出与当前文件名匹配且优先级更高的文件

        higher_priority_file_exists = False
        for higher_ext in priority_order[:priority_order.index(ext)]:
            if os.path.exists(os.path.join(folder_path, base_name + higher_ext)):
                higher_priority_file_exists = True
                print(f"Skipping file: {file_name} (higher priority file {base_name + higher_ext} exists)")
                break

        # 如果存在更高优先级的文件，跳过当前文件
        if higher_priority_file_exists:
            continue

        file_path = os.path.join(folder_path, file_name)

        # 处理文件，如果匹配到平台
        if platform_name in platform_mapping.keys():
            # 查找对应平台的配置
            platform_config = next((p for p in config['platforms'] if p['platform_name'] == platform_name), None)
            if platform_config is None:
                result_log.append(f"{file_name} skipped (platform config not found).")
                continue

            print(f"Processing file: {file_name}")
            success = process_inventory_file(file_path, config, sku_mapping_df, shared_inventory_df,
                                             realtime_inventory_df, warehouse_mapping_df, platform_mapping,
                                             output_dir, result_log, platform_config, output_temp_dir)
            if success:
                # 标记该基础名称的文件已处理，跳过其他同名文件
                processed_files.add(base_name)
            else:
                all_files_processed_successfully = False
        else:
            result_log.append(f"{file_name} skipped (not matching any platform).")

    # 创建日志文件，记录结果
    current_time = datetime.now().strftime("%H点%M分")
    log_filename = f'已处理完成{current_time}.txt' if all_files_processed_successfully else f'处理存在错误{current_time}.txt'
    with open(os.path.join(output_dir, log_filename), 'w', encoding='utf-8') as log_file:
        log_file.write("\n".join(result_log))


def main():
    current_date = datetime.now().strftime("%m月%d日")
    current_date_realtime = datetime.now().strftime("%Y%m%d")
    config_path = r"D:\my-project\inventory_update\Inventory_updates_wei\区分仓库_project\config-9月13日_区分仓库.json"
    sku_mapping_path = r"W:\库存查询原始表\配置文件\SKU映射汇总表_加来源test.csv"
    shared_inventory_path = rf"W:\库存查询原始表\每日库存数据\504-shared_inventory-{current_date}.csv"
    print(shared_inventory_path)
    # source_folder = r"W:\库存查询原始表\RPA_魏土金_test\原始表"
    source_folder = r"W:\库存查询原始表"
    # output_base_folder = r"W:\库存查询原始表\RPA_魏土金_test\结果表"
    output_base_folder = r"W:\库存查询原始表\RPA_魏土金_test\结果表"
    realtime_inventory_path = rf"W:\库存查询原始表\每日库存数据\海外仓库存表{current_date_realtime}.xls"
    warehouse_mapping_path = r"W:\库存查询原始表\配置文件\仓库映射.xlsx"
    output_temp_dir = r"W:\库存查询原始表\RPA_魏土金_test\output_dir"


    config = load_config(config_path)

    platform_mapping = {
        "SEARS": "Sears",
        "Allegro": "Allegro",
        "AE": "速卖通",
        "Wal": "Walmart",
        "MC": "MoreCommerce",
        "HZ": "Houzz",
        "Shein": "Shein",
        "NEWEGG": "NewEgg",
        "Shopify": "Shopify",
        "TEMU": "TEMU",
        "ebay": "ebay",
        "JD": "京东国际平台",
        # ---- 区分仓库的平台 -----
        "HD": "HomeDepot",
        "LW": "Lowes",
        "WF": "Wayfair",
        "OS": "OverStock",
        "Target": "Target",
        "AMZ": "Amazon",
        "TK": "TikTok",
    }

    # 加载SKU映射表
    sku_mapping_df = None
    try:
        # 尝试使用主要编码读取CSV文件
        sku_mapping_df = pd.read_csv(sku_mapping_path, encoding='gbk', dtype={"平台ID": str, "渠道sku": str})
    except UnicodeDecodeError:
        try:
            sku_mapping_df = pd.read_csv(sku_mapping_path, encoding='utf-8-sig')
        except Exception as e:
            raise e
    print("DataFrame Columns:", sku_mapping_df.columns.tolist())

    # 在程序开始时读取数据

    # 将索引列转换为字符串类型（如果需要）
    sku_mapping_df['渠道sku'] = sku_mapping_df['渠道sku'].astype(str)
    sku_mapping_df['平台'] = sku_mapping_df['平台'].astype(str)
    sku_mapping_df['来源'] = sku_mapping_df['来源'].astype(str)

    # # 设置 MultiIndex
    sku_mapping_df.set_index(['渠道sku', '平台', '来源'], inplace=True, drop=False)
    sku_mapping_df.sort_index(inplace=True)
    print("DataFrame Columns:", sku_mapping_df.columns.tolist())

    # 加载共享库存表
    shared_inventory_df = pd.read_csv(shared_inventory_path, encoding='gbk')
    print(f"共享库存表已加载，共有 {len(shared_inventory_df)} 条记录。")

    # 加载实时库存表
    realtime_inventory_path = convert_xls_to_xlsx(realtime_inventory_path, output_temp_dir)
    realtime_inventory_df = pd.read_excel(realtime_inventory_path, engine='openpyxl')
    print(f"实时库存表已加载，共有 {len(realtime_inventory_df)} 条记录。")

    # 加载仓库映射表
    warehouse_mapping_df = pd.read_excel(warehouse_mapping_path, engine='openpyxl', dtype={'平台仓库ID': str})
    print(f"仓库映射表已加载，共有 {len(warehouse_mapping_df)} 条记录。")

    current_date = datetime.now().strftime("%Y-%m-%d")
    output_folder = os.path.join(output_base_folder, current_date)
    os.makedirs(output_folder, exist_ok=True)

    # 文件夹列表 - 指定需要跳过的文件夹。同时记录处理到哪个文件夹了
    skip_folders = []

    operation_record_path = os.path.join(output_folder, '程序处理时长记录.txt')
    with open(operation_record_path, 'a', encoding='utf-8') as file:
        # 获取当前的日期和时间
        current_datetime = datetime.now()
        # 写入字符串到文件
        formatted_datetime = current_datetime.strftime("%m月%d日 %H:%M:%S" + "\n")
        file.write(formatted_datetime)

    for sub_folder_name in os.listdir(source_folder):
        # 检查当前子文件夹是否在 skip_folders 列表中，若在则跳过
        if sub_folder_name in skip_folders:
            print(f"Skipping folder: {sub_folder_name}")
            continue

        sub_folder_path = os.path.join(source_folder, sub_folder_name)
        # ["组", "站", "院"]  ["研究院-易锦涛"]  "先锋组-陈天", "新平台组-陈成", "战狼B组-黄桂璇"
        # if os.path.isdir(sub_folder_path) and any(keyword in sub_folder_name for keyword in ["组", "站", "院"]):
        # if os.path.isdir(sub_folder_path) and any(keyword in sub_folder_name for keyword in
        #                                           ["研究院-易锦涛"]):
        if os.path.isdir(sub_folder_path) and any(keyword in sub_folder_name for keyword in ["新平台组-陈成"]):
            print(f"Processing folder: {sub_folder_name}")
            skip_folders.append(sub_folder_name)
            print(f"已处理文件夹: {skip_folders}")
            output_sub_folder = os.path.join(output_folder, sub_folder_name)
            os.makedirs(output_sub_folder, exist_ok=True)
            # 记录开始时间
            start_time = time.perf_counter()
            process_folder(sub_folder_path, config, sku_mapping_df, shared_inventory_df, realtime_inventory_df,
                           warehouse_mapping_df, platform_mapping, output_sub_folder, output_temp_dir)
            # 记录结束时间
            end_time = time.perf_counter()
            # 计算运行时间
            elapsed_time = end_time - start_time
            # 将秒数转换成分钟和秒
            minutes, seconds = divmod(int(elapsed_time), 60)
            record = f"{sub_folder_name}:  运行时长{minutes}分{seconds}秒"

            # 打开文件，以追加模式（'a'）打开
            with open(operation_record_path, 'a', encoding='utf-8') as file:
                # 写入字符串到文件
                file.write(record + "\n")


if __name__ == "__main__":
    main()
